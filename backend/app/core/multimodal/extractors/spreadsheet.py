# (c) 2026 AgentFlow-Eval
"""Spreadsheet / table extraction (CSV stdlib, XLSX openpyxl/pandas)."""

from __future__ import annotations

import csv
import io
import logging
from pathlib import Path
from typing import Any

from app.core.multimodal.types import ExtractResult, MediaKind

logger = logging.getLogger(__name__)


def extract_spreadsheet(data: bytes, filename: str = "data.xlsx") -> ExtractResult:
    """Extract tabular data as text + structured row samples."""
    ext = Path(filename).suffix.lower()
    warnings: list[str] = []
    tables: list[dict[str, Any]] = []
    text_parts: list[str] = []
    features: dict[str, Any] = {"byte_size": len(data), "extension": ext}
    metadata: dict[str, Any] = {"filename": filename}

    if ext in {".csv", ".tsv", ".txt"}:
        delimiter = "\t" if ext == ".tsv" else ","
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = data.decode("latin-1", errors="replace")
            warnings.append("decoded with latin-1 fallback")
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        tables.append(_table_from_rows("sheet1", rows))
        text_parts.append(_rows_to_markdown("sheet1", rows))
        features["sheet_count"] = 1
        features["row_count"] = max(0, len(rows) - 1) if rows else 0
    elif ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        result = _extract_xlsx(data)
        tables.extend(result["tables"])
        text_parts.extend(result["text_parts"])
        warnings.extend(result["warnings"])
        features.update(result["features"])
    elif ext == ".xls":
        # Optional xlrd path
        try:
            import pandas as pd  # type: ignore

            df = pd.read_excel(io.BytesIO(data), sheet_name=None, engine=None)
            features["sheet_count"] = len(df)
            total_rows = 0
            for name, frame in df.items():
                rows = [list(frame.columns.astype(str))] + frame.astype(
                    str
                ).values.tolist()
                tables.append(_table_from_rows(str(name), rows))
                text_parts.append(_rows_to_markdown(str(name), rows))
                total_rows += max(0, len(rows) - 1)
            features["row_count"] = total_rows
        except Exception as exc:
            warnings.append(f"xls extract failed (install pandas/xlrd): {exc}")
    else:
        warnings.append(f"unsupported spreadsheet extension: {ext}")

    return ExtractResult(
        kind=MediaKind.SPREADSHEET,
        text="\n\n".join(text_parts),
        features=features,
        metadata=metadata,
        tables=tables,
        warnings=warnings,
    )


def _extract_xlsx(data: bytes) -> dict[str, Any]:
    warnings: list[str] = []
    tables: list[dict[str, Any]] = []
    text_parts: list[str] = []
    features: dict[str, Any] = {}

    # Prefer openpyxl (lighter); pandas as fallback
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        features["sheet_count"] = len(wb.sheetnames)
        total_rows = 0
        for name in wb.sheetnames:
            ws = wb[name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append(["" if c is None else str(c) for c in row])
            tables.append(_table_from_rows(name, rows))
            text_parts.append(_rows_to_markdown(name, rows))
            total_rows += max(0, len(rows) - 1)
        features["row_count"] = total_rows
        wb.close()
        return {
            "tables": tables,
            "text_parts": text_parts,
            "warnings": warnings,
            "features": features,
        }
    except ImportError:
        warnings.append("openpyxl not installed; trying pandas")
    except Exception as exc:
        warnings.append(f"openpyxl failed: {exc}")

    try:
        import pandas as pd  # type: ignore

        book = pd.read_excel(io.BytesIO(data), sheet_name=None, engine="openpyxl")
        features["sheet_count"] = len(book)
        total_rows = 0
        for name, frame in book.items():
            rows = [list(map(str, frame.columns))] + frame.astype(str).values.tolist()
            tables.append(_table_from_rows(str(name), rows))
            text_parts.append(_rows_to_markdown(str(name), rows))
            total_rows += max(0, len(rows) - 1)
        features["row_count"] = total_rows
    except Exception as exc:
        warnings.append(f"pandas excel extract failed: {exc}")

    return {
        "tables": tables,
        "text_parts": text_parts,
        "warnings": warnings,
        "features": features,
    }


def _table_from_rows(name: str, rows: list[list[str]]) -> dict[str, Any]:
    header = rows[0] if rows else []
    body = rows[1:51] if len(rows) > 1 else []  # cap sample
    return {
        "name": name,
        "columns": header,
        "row_count": max(0, len(rows) - 1) if rows else 0,
        "sample_rows": body,
    }


def _rows_to_markdown(name: str, rows: list[list[str]], max_rows: int = 30) -> str:
    if not rows:
        return f"### {name}\n(empty)"
    lines = [f"### {name}"]
    header = rows[0]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join("---" for _ in header) + " |")
    for row in rows[1 : max_rows + 1]:
        # pad/truncate to header width
        cells = list(row) + [""] * max(0, len(header) - len(row))
        lines.append("| " + " | ".join(cells[: len(header)]) + " |")
    if len(rows) - 1 > max_rows:
        lines.append(f"... ({len(rows) - 1 - max_rows} more rows)")
    return "\n".join(lines)
